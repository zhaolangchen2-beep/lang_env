#!/usr/bin/env python3
"""
分析指定的 pyperformance JSON 结果文件，
生成 Excel 对比表格和分页性能趋势图（每20个用例一张图）。

用法：
  python get_stat.py a.json d.json c.json b.json       # 完整输出（打屏+Excel+图）
  python get_stat.py -c a.json d.json c.json b.json     # 仅打屏，不输出文件
  python get_stat.py                                     # 自动扫描当前目录

依赖：pip install openpyxl matplotlib
"""

import json
import glob
import os
import sys
import argparse
from math import exp, log

import matplotlib
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

matplotlib.rcParams['font.sans-serif'] = [
    'SimHei', 'WenQuanYi Micro Hei', 'Noto Sans CJK SC',
    'Microsoft YaHei', 'DejaVu Sans',
]
matplotlib.rcParams['axes.unicode_minus'] = False


# ═══════════════════════════════════════════════════════════
#  工具函数
# ═══════════════════════════════════════════════════════════

def load_benchmark_data(json_file):
    """从 pyperformance JSON 中提取各用例平均耗时(秒)。"""
    with open(json_file, 'r', encoding='utf-8') as f:
        data = json.load(f)

    results = {}
    if 'benchmarks' in data:
        for bench in data['benchmarks']:
            name = None
            if 'metadata' in bench and 'name' in bench['metadata']:
                name = bench['metadata']['name']
            if name is None:
                for run in bench.get('runs', []):
                    if 'metadata' in run and 'name' in run['metadata']:
                        name = run['metadata']['name']
                        break
            if name is None:
                continue
            values = []
            for run in bench.get('runs', []):
                if 'values' in run:
                    values.extend(run['values'])
            if values:
                results[name] = sum(values) / len(values)
    elif 'runs' in data:
        name = data.get('metadata', {}).get('name', 'unknown')
        values = []
        for run in data['runs']:
            if 'values' in run:
                values.extend(run['values'])
        if values:
            results[name] = sum(values) / len(values)
    return results


def choose_unit(value_seconds):
    """根据数值量级选取合适的显示单位。"""
    if value_seconds >= 1.0:
        return 'sec', 1.0
    elif value_seconds >= 1e-3:
        return 'ms', 1e-3
    elif value_seconds >= 1e-6:
        return 'us', 1e-6
    else:
        return 'ns', 1e-9


def geometric_mean(values):
    """计算一组正数的几何平均值。"""
    pos = [v for v in values if v > 0]
    if not pos:
        return 1.0
    return exp(sum(log(v) for v in pos) / len(pos))


# ═══════════════════════════════════════════════════════════
#  Excel 输出
# ═══════════════════════════════════════════════════════════

def save_excel(common_benchmarks, valid_files, all_data, units, divisors,
               perf_geo_means, xlsx_path='benchmark_comparison.xlsx'):
    """生成带格式的 Excel 对比表格。"""
    wb = Workbook()
    ws = wb.active
    ws.title = '性能对比'

    file_labels = [os.path.basename(f) for f in valid_files]

    header_font = Font(bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    summary_font = Font(bold=True, size=11, color='FFFFFF')
    summary_fill = PatternFill(start_color='E26B0A', end_color='E26B0A', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )
    center = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')

    headers = ['Benchmark', 'Unit'] + file_labels
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border

    for row_idx, bench in enumerate(common_benchmarks, 2):
        c = ws.cell(row=row_idx, column=1, value=bench)
        c.alignment = left_align
        c.border = thin_border

        c = ws.cell(row=row_idx, column=2, value=units[bench])
        c.alignment = center
        c.border = thin_border

        for col_off, jf in enumerate(valid_files):
            v = all_data[jf][bench] / divisors[bench]
            c = ws.cell(row=row_idx, column=3 + col_off, value=round(v, 4))
            c.number_format = '0.0000'
            c.alignment = center
            c.border = thin_border

    summary_row = len(common_benchmarks) + 2
    c = ws.cell(row=summary_row, column=1, value='性能对比')
    c.font = summary_font
    c.fill = summary_fill
    c.alignment = left_align
    c.border = thin_border

    c = ws.cell(row=summary_row, column=2, value='NA')
    c.font = summary_font
    c.fill = summary_fill
    c.alignment = center
    c.border = thin_border

    for col_off, gm in enumerate(perf_geo_means):
        c = ws.cell(row=summary_row, column=3 + col_off, value=round(gm, 4))
        c.font = summary_font
        c.fill = summary_fill
        c.number_format = '0.0000'
        c.alignment = center
        c.border = thin_border

    ws.column_dimensions['A'].width = max(len(b) for b in common_benchmarks) + 4
    ws.column_dimensions['B'].width = 8
    for col_off, lb in enumerate(file_labels):
        col_letter = chr(ord('C') + col_off)
        ws.column_dimensions[col_letter].width = max(len(lb), 12) + 4

    ws.freeze_panes = 'A2'
    wb.save(xlsx_path)
    print(f"📄  Excel 表格已保存到 {xlsx_path}")


# ═══════════════════════════════════════════════════════════
#  绘图（分页）
# ═══════════════════════════════════════════════════════════

def plot_trends_paginated(common_benchmarks, valid_files, perf_ratios, perf_geo_means,
                          benchmarks_per_page=20, base_name='benchmark_trends'):
    file_labels = [os.path.basename(f) for f in valid_files]
    n_bench = len(common_benchmarks)
    n_pages = (n_bench + benchmarks_per_page - 1) // benchmarks_per_page
    x = list(range(len(valid_files)))

    print(f"\n📊  共 {n_bench} 个用例，将生成 {n_pages} 张图（每页最多 {benchmarks_per_page} 个用例）")

    for page_idx in range(n_pages):
        start = page_idx * benchmarks_per_page
        end = min(start + benchmarks_per_page, n_bench)
        page_benchmarks = common_benchmarks[start:end]
        n_plots = len(page_benchmarks) + 1

        subplot_h = 2.0
        fig_h = n_plots * subplot_h
        fig_w = max(10, len(valid_files) * 1.5)

        fig, axes = plt.subplots(
            n_plots, 1,
            figsize=(fig_w, fig_h),
            sharex=True,
        )
        if n_plots == 1:
            axes = [axes]

        for idx, bench in enumerate(page_benchmarks):
            ax = axes[idx]
            y = [perf_ratios[jf][bench] for jf in valid_files]
            ax.plot(x, y, 'o-', color='steelblue', linewidth=1.8, markersize=6)
            ax.fill_between(x, 1.0, y, alpha=0.10, color='steelblue')
            ax.axhline(y=1.0, color='red', linestyle='--', alpha=0.5, linewidth=0.8)

            for i, v in enumerate(y):
                ax.annotate(f'{v:.3f}', (i, v),
                            textcoords='offset points', xytext=(0, 8),
                            ha='center', fontsize=7, color='steelblue')

            ax.set_ylabel('Ratio', fontsize=8)
            ax.set_title(bench, fontsize=10, fontweight='bold', loc='left', pad=4)
            ax.grid(True, alpha=0.2)
            ax.tick_params(labelsize=8)

        ax = axes[-1]
        color = 'forestgreen' if perf_geo_means[-1] >= 1.0 else 'orangered'
        ax.plot(x, perf_geo_means, 'o-', color=color, linewidth=2.8, markersize=9)
        ax.fill_between(x, 1.0, perf_geo_means, alpha=0.18, color=color)
        ax.axhline(y=1.0, color='red', linestyle='--', alpha=0.5, linewidth=0.8)

        for i, gm in enumerate(perf_geo_means):
            ax.annotate(f'{gm:.4f}', (i, gm),
                        textcoords='offset points', xytext=(0, 12),
                        ha='center', fontsize=10, fontweight='bold', color=color)

        ax.set_ylabel('Geo Mean', fontsize=9)
        ax.set_title('★ Overall Performance (Geometric Mean)', fontsize=11,
                     fontweight='bold', loc='left', color=color, pad=4)
        ax.grid(True, alpha=0.2)

        ax.set_xticks(x)
        ax.set_xticklabels(file_labels, rotation=45, ha='right', fontsize=9)
        ax.set_xlabel('JSON Files', fontsize=10)

        png_path = f"{base_name}_part{page_idx + 1}.png"
        plt.tight_layout(h_pad=0.8)
        plt.savefig(png_path, dpi=150, bbox_inches='tight')
        print(f"   ✅ {png_path}  ({len(page_benchmarks)} 个用例)")
        plt.close()

    print()


# ═══════════════════════════════════════════════════════════
#  参数解析
# ═══════════════════════════════════════════════════════════

def parse_args():
    parser = argparse.ArgumentParser(
        description='pyperformance 多版本性能对比分析工具',
        epilog='示例:\n'
               '  python get_stat.py a.json d.json c.json b.json\n'
               '  python get_stat.py -c a.json d.json c.json b.json\n'
               '  python get_stat.py\n',
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument(
        '-c', '--console-only',
        action='store_true',
        help='仅打屏输出对比数据，不生成 Excel 和趋势图',
    )
    parser.add_argument(
        'json_files',
        nargs='*',
        help='要对比的 JSON 文件（按指定顺序）；不传则自动扫描当前目录',
    )
    return parser.parse_args()


# ═══════════════════════════════════════════════════════════
#  主流程
# ═══════════════════════════════════════════════════════════

def main():
    args = parse_args()

    # ── 1. 获取文件列表 ──
    if args.json_files:
        json_files = args.json_files
        for jf in json_files:
            if not os.path.isfile(jf):
                print(f"❌  文件不存在: {jf}")
                sys.exit(1)
    else:
        json_files = sorted(glob.glob('*.json'))

    if not json_files:
        print("❌  未找到 JSON 文件。")
        print(f"用法: python {os.path.basename(__file__)} [-c] a.json d.json c.json b.json")
        sys.exit(1)

    print(f"📂  共 {len(json_files)} 个 JSON 文件（按指定顺序）：")
    all_data = {}
    for jf in json_files:
        try:
            all_data[jf] = load_benchmark_data(jf)
            print(f"   ✅ {jf}  ({len(all_data[jf])} 个用例)")
        except Exception as e:
            print(f"   ⚠️  {jf} 加载失败: {e}")

    valid_files = [jf for jf in json_files if jf in all_data and all_data[jf]]
    if len(valid_files) < 1:
        print("❌  没有有效数据。")
        sys.exit(1)

    # ── 2. 取公共用例 ──
    common = set(all_data[valid_files[0]].keys())
    for jf in valid_files[1:]:
        common &= set(all_data[jf].keys())
    common_benchmarks = sorted(common)

    if not common_benchmarks:
        print("❌  多个文件之间没有公共用例。")
        sys.exit(1)
    print(f"\n🔗  公共用例数量: {len(common_benchmarks)}\n")

    # ── 3. 确定显示单位 ──
    units = {}
    divisors = {}
    for bench in common_benchmarks:
        u, d = choose_unit(all_data[valid_files[0]][bench])
        units[bench] = u
        divisors[bench] = d

    # ── 4. 计算性能对比 ──
    perf_ratios = {}
    perf_geo_means = []
    for jf in valid_files:
        ratios = {}
        for bench in common_benchmarks:
            ratios[bench] = all_data[valid_files[0]][bench] / all_data[jf][bench]
        perf_ratios[jf] = ratios
        perf_geo_means.append(geometric_mean(list(ratios.values())))

    # ── 5. 终端表格（始终输出） ──
    file_labels = [os.path.basename(f) for f in valid_files]
    col_w = max(max(len(l) for l in file_labels), 12)
    name_w = max(max(len(b) for b in common_benchmarks), 8)
    unit_w = 5

    header = f"{'Benchmark':<{name_w}}  {'Unit':<{unit_w}}"
    for lb in file_labels:
        header += f"  {lb:>{col_w}}"
    sep = '=' * len(header)
    thin = '-' * len(header)

    print(sep)
    print(header)
    print(sep)
    for bench in common_benchmarks:
        row = f"{bench:<{name_w}}  {units[bench]:<{unit_w}}"
        for jf in valid_files:
            v = all_data[jf][bench] / divisors[bench]
            row += f"  {v:>{col_w}.4f}"
        print(row)
    print(thin)
    row = f"{'性能对比':<{name_w}}  {'NA':<{unit_w}}"
    for gm in perf_geo_means:
        row += f"  {gm:>{col_w}.4f}"
    print(row)
    print(sep)

    # ── 6. 如果指定了 -c，到此结束 ──
    if args.console_only:
        print("\n💡  已指定 -c 参数，仅打屏输出，跳过 Excel 和趋势图生成。")
        return

    # ── 7. 保存 Excel ──
    save_excel(common_benchmarks, valid_files, all_data, units, divisors,
               perf_geo_means)

    # ── 8. 分页绘图 ──
    plot_trends_paginated(common_benchmarks, valid_files, perf_ratios, perf_geo_means,
                          benchmarks_per_page=20)


if __name__ == '__main__':
    main()